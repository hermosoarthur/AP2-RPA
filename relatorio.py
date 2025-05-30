from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sqlite3
from datetime import datetime

def obter_dados_paises():
    conn = sqlite3.connect('paises.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM paises")
    dados = cursor.fetchall()
    conn.close()
    return dados

def obter_dados_livros():
    conn = sqlite3.connect('livraria.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM livros")
    dados = cursor.fetchall()
    conn.close()
    return dados

def formatar_celula(celula):
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    celula.border = borda
    celula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def gerar_relatorio_excel(nome_aluno):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório RPA"

        cabecalho_principal = ws.cell(row=1, column=1, value=f"Relatório de Dados - {nome_aluno}")
        cabecalho_principal.font = Font(bold=True, size=16)
        ws.merge_cells('A1:N1')
        
        data_geracao = ws.cell(row=2, column=1, value=f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        data_geracao.font = Font(italic=True)
        ws.merge_cells('A2:N2')

        linha_atual = 4
        ws.cell(row=linha_atual, column=1, value="DADOS DOS PAÍSES").font = Font(bold=True, size=14)
        ws.merge_cells(f'A{linha_atual}:N{linha_atual}')
        linha_atual += 1
        
        cabecalhos_paises = [
            "Nome Comum", "Nome Oficial", "Capital", "Continente",
            "Região", "Sub-região", "População", "Área (km²)",
            "Moeda (Nome)", "Moeda (Símbolo)", "Idioma Principal",
            "Fuso Horário", "URL da Bandeira", "Data de Inserção"
        ]
        
        for col, cabecalho in enumerate(cabecalhos_paises, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=cabecalho)
            celula.font = Font(bold=True)
            formatar_celula(celula)
        
        linha_atual += 1
        
        paises = obter_dados_paises()
        for pais in paises:
            for col, valor in enumerate(pais[1:], start=1): 
                celula = ws.cell(row=linha_atual, column=col, value=valor)
                formatar_celula(celula)
            linha_atual += 1
        
        linha_atual += 2  
        ws.cell(row=linha_atual, column=1, value="DADOS DOS LIVROS").font = Font(bold=True, size=14)
        ws.merge_cells(f'A{linha_atual}:N{linha_atual}')
        linha_atual += 1
        
        cabecalhos_livros = [
            "Título", "Preço (£)", "Avaliação", "Disponibilidade", "Data de Inserção"
        ]
        
        for col, cabecalho in enumerate(cabecalhos_livros, start=1):
            celula = ws.cell(row=linha_atual, column=col, value=cabecalho)
            celula.font = Font(bold=True)
            formatar_celula(celula)
            if col > len(cabecalhos_livros):
                break
        
        linha_atual += 1
        
        livros = obter_dados_livros()
        for livro in livros:
            for col, valor in enumerate(livro[1:], start=1):
                celula = ws.cell(row=linha_atual, column=col, value=valor)
                formatar_celula(celula)
                if col > len(cabecalhos_livros):
                    break
            linha_atual += 1
        
        for i, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col_cells:

                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        nome_arquivo = f"relatorio_rpa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(nome_arquivo)
        print(f"\n✅ Relatório gerado com sucesso: {nome_arquivo}")
        return nome_arquivo
    except Exception as e:
        print(f"\n❌ Erro ao gerar relatório: {str(e)}")
        return None

def executar_parte_3():
    """Função principal da Parte 3"""
    print("\n" + "="*50)
    print("GERAÇÃO DE RELATÓRIO")
    print("="*50)
    
    nome_aluno = input("Digite seu nome: ").strip()
    gerar_relatorio_excel(nome_aluno)